from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem, Profile
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from rest_framework import viewsets, permissions
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated, SAFE_METHODS
from .serializers import (
    ProductSerializer, CategorySerializer, ManufacturerSerializer,
    CartSerializer, CartItemSerializer, ProfileSerializer
)
from .models import Order, OrderItem
from .serializers import OrderSerializer
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm

def index(request):
    """Главная страница"""
    products = Product.objects.all().order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'popular_products': products,
        'categories': categories,
    })


def catalog(request):
    products = Product.objects.all().order_by('id')
    
    # Получаем параметры из GET
    category_id = request.GET.get('category')
    manufacturer_id = request.GET.get('manufacturer')
    search_query = request.GET.get('search', '').strip()
    
    # ПРИМЕНЯЕМ ФИЛЬТРЫ
    if category_id and category_id.isdigit():
        products = products.filter(category_id=int(category_id))
    
    if manufacturer_id and manufacturer_id.isdigit():
        products = products.filter(manufacturer_id=int(manufacturer_id))
    
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Контекст для шаблона
    context = {
        'products': page_obj,  # ← сюда должны попасть ОТФИЛЬТРОВАННЫЕ товары
        'categories': Category.objects.all(),
        'manufacturers': Manufacturer.objects.all(),
        'selected_category': int(category_id) if category_id and category_id.isdigit() else None,
        'selected_manufacturer': int(manufacturer_id) if manufacturer_id and manufacturer_id.isdigit() else None,
        'search_query': search_query,
    }
    
    # ОТЛАДКА (вывод в консоль)
    print(f"=== ФИЛЬТРЫ ===")
    print(f"Категория: {category_id}")
    print(f"Производитель: {manufacturer_id}")
    print(f"Поиск: {search_query}")
    print(f"Найдено товаров: {products.count()}")
    print(f"=================")
    
    return render(request, 'shop/catalog.html', context)

def product_detail(request, pk):
    """Детальная информация о товаре"""
    product = get_object_or_404(Product, id=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


def cart_view(request):
    """Просмотр корзины (доступно всем)"""
    cart_items = []
    total_price = 0
    
    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user)
        cart_items = cart.items.all()
        total_price = cart.total_price()
    
    context = {
        'cart_items': cart_items,
        'total_price': total_price,
    }
    return render(request, 'shop/cart.html', context)


@login_required
def add_to_cart(request, product_id):
    """Добавление товара в корзину (только для авторизованных)"""
    product = get_object_or_404(Product, id=product_id)
    
    cart, created = Cart.objects.get_or_create(user=request.user)
    
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': 1}
    )
    
    if not created:
        if cart_item.quantity + 1 <= product.quantity:
            cart_item.quantity += 1
            cart_item.save()
            messages.success(request, f'Товар "{product.name}" добавлен в корзину (+1)')
        else:
            messages.error(request, f'Недостаточно товара на складе. Доступно: {product.quantity}')
    else:
        messages.success(request, f'Товар "{product.name}" добавлен в корзину')
    
    return redirect('catalog')


@login_required
def update_cart(request, item_id):
    """Обновление количества товара в корзине"""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    
    if request.method == 'POST':
        new_quantity = int(request.POST.get('quantity', 1))
        
        if new_quantity <= 0:
            cart_item.delete()
            messages.success(request, f'Товар "{cart_item.product.name}" удалён из корзины')
        elif new_quantity <= cart_item.product.quantity:
            cart_item.quantity = new_quantity
            cart_item.save()
            messages.success(request, f'Количество товара "{cart_item.product.name}" обновлено')
        else:
            messages.error(request, f'Недостаточно товара на складе. Доступно: {cart_item.product.quantity}')
    
    return redirect('cart_view')


@login_required
def remove_from_cart(request, item_id):
    """Удаление товара из корзины"""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'Товар "{product_name}" удалён из корзины')
    return redirect('cart_view')


def user_login(request):
    """Вход по email"""
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        try:
            user = User.objects.get(email=email)
            user = authenticate(request, username=user.username, password=password)
            if user is not None:
                login(request, user)
                return redirect('catalog')
            else:
                messages.error(request, 'Неверный пароль')
        except User.DoesNotExist:
            messages.error(request, 'Пользователь с таким email не найден')
    
    return render(request, 'shop/login.html')


def register(request):
    """Регистрация пользователя"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        full_name = request.POST.get('full_name', '')
        phone = request.POST.get('phone', '')
        
        if password1 != password2:
            messages.error(request, 'Пароли не совпадают')
            return render(request, 'shop/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким логином уже существует')
            return render(request, 'shop/register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Пользователь с таким email уже существует')
            return render(request, 'shop/register.html')
        
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password1
        )
        
        profile = user.profile
        profile.full_name = full_name
        profile.phone = phone
        profile.save()
        
        login(request, user)
        messages.success(request, 'Регистрация прошла успешно!')
        return redirect('catalog')
    
    return render(request, 'shop/register.html')


def user_logout(request):
    logout(request)
    return redirect('catalog')


@login_required
def profile_view(request):
    return render(request, 'shop/profile.html')


def generate_excel_receipt(order, cart_items, total_price):
    """Генерация Excel-файла чека"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чек"
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "ЧЕК О ФОРМЛЕНИИ ЗАКАЗА"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Номер заказа: #{order.id}"
    ws['A4'] = f"Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}"
    ws['A5'] = f"Покупатель: {order.user.username}"
    ws['A6'] = f"Адрес доставки: {order.address}"
    ws['A7'] = f"Телефон: {order.phone}"
    
    headers = ['№', 'Товар', 'Категория', 'Цена', 'Кол-во', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.border = border
    
    row = 10
    for idx, item in enumerate(cart_items, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.product.name).border = border
        ws.cell(row=row, column=3, value=item.product.category.name).border = border
        ws.cell(row=row, column=4, value=float(item.product.price)).border = border
        ws.cell(row=row, column=5, value=item.quantity).border = border
        ws.cell(row=row, column=6, value=float(item.element_price())).border = border
        row += 1
    
    ws.cell(row=row, column=5, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_price)).font = Font(bold=True)
    
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all()
    total_price = cart.total_price()
    
    if request.method == 'POST':
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        
        if not address or not phone:
            messages.error(request, 'Заполните все поля')
            return render(request, 'shop/checkout.html', {
                'cart_items': cart_items,
                'total_price': total_price,
            })
        
        order = Order.objects.create(
            user=request.user,
            address=address,
            phone=phone,
            total_price=total_price
        )
        
        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )
            item.product.quantity -= item.quantity
            item.product.save()
        
        excel_file = generate_excel_receipt(order, cart_items, total_price)
        
        subject = f'Чек на заказ #{order.id} в магазине рюкзаков'
        message = f"""
Здравствуйте, {request.user.username}!

Ваш заказ #{order.id} успешно оформлен.

Детали заказа:
- Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}
- Адрес доставки: {address}
- Телефон: {phone}
- Итого: {total_price} руб.

Чек во вложении.

Спасибо за покупку!
"""
        
        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL or 'noreply@shop.com',
                [request.user.email],
                fail_silently=False,
            )
            messages.success(request, f'Заказ #{order.id} оформлен! Чек отправлен на {request.user.email}')
        except Exception as e:
            messages.warning(request, f'Заказ оформлен, но письмо не отправлено. Ошибка: {str(e)}')
        
        cart_items.delete()
        return redirect('catalog')
    
    return render(request, 'shop/checkout.html', {
        'cart_items': cart_items,
        'total_price': total_price,
    })


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().order_by('id')
    serializer_class = ProductSerializer
    
    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get('category')
        manufacturer_id = self.request.query_params.get('manufacturer')
        
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        if manufacturer_id:
            queryset = queryset.filter(manufacturer_id=manufacturer_id)
        
        return queryset


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAuthenticated]


class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.none()
    serializer_class = CartSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Cart.objects.filter(user=self.request.user)
    
    @action(detail=False, methods=['get'])
    def my_cart(self, request):
        cart, created = Cart.objects.get_or_create(user=request.user)
        serializer = self.get_serializer(cart)
        return Response(serializer.data)


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.none()
    serializer_class = CartItemSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return CartItem.objects.filter(cart__user=self.request.user)
    
    def perform_create(self, serializer):
        cart, created = Cart.objects.get_or_create(user=self.request.user)
        serializer.save(cart=cart)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def api_me(request):
    """Получение и обновление профиля"""
    profile = request.user.profile
    
    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)
    
    elif request.method == 'PATCH':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    


class OrderViewSet(viewsets.ReadOnlyModelViewSet):
    """API для заказов"""
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser or user.profile.role == 'ADMIN':
            return Order.objects.all()
        return Order.objects.filter(user=user)
    
@login_required
def settings_view(request):
    """Страница настроек (смена пароля)"""
    
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Обновляем сессию, чтобы пользователь не вылетел
            update_session_auth_hash(request, user)
            messages.success(request, 'Пароль успешно изменён!')
            return redirect('settings')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'shop/settings.html', {
        'form': form,
    })